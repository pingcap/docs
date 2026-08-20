from __future__ import annotations

import importlib
import json
import sys
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


SCRIPTS_DIR = Path(__file__).resolve().parents[2]
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

ai_client = importlib.import_module("release-notes-ai-generator.ai_client")
excel_workbook = importlib.import_module("release-notes-ai-generator.excel_workbook")
models = importlib.import_module("release-notes-ai-generator.models")


class FakeGitHubClient:
    def get_issue(self, issue_url: str):
        return models.IssueInfo(
            url=issue_url,
            title="Issue title",
            body="Issue body",
            labels=[],
        )

    def get_pull(self, pr_url: str):
        return models.PullInfo(
            url=pr_url,
            title="PR title",
            body="PR body",
            author="contributor",
            head_ref="feature",
            base_ref="master",
            files_summary="file: config.go\npatch:\n+ tidb_example = ON",
        )


class FakeAIClient:
    def __init__(self, results):
        self.results = list(results)
        self.calls = 0
        self.prompts = []
        self.task_flags = []

    def generate(
        self,
        prompt: str,
        expected_links: list[str],
        contributors: list[str],
        expected_pr_links: list[str],
        generate_release_note: bool = True,
        generate_doc_impact: bool = True,
    ):
        self.calls += 1
        self.prompts.append(prompt)
        self.task_flags.append((generate_release_note, generate_doc_impact))
        return self.results.pop(0)


class RecordingAIClient(ai_client.AIClient):
    def __init__(self, output):
        self.output = output
        self.schemas = []

    def _run(self, prompt, output_schema):
        self.schemas.append(output_schema)
        return json.dumps(self.output)


def detected_impact(pr_url: str):
    return models.VariableOrConfigDocImpact(
        status="detected",
        changes=[
            models.DocImpactChange(
                kind="system_variable",
                name="tidb_example",
                change_type="modified",
                description="Changes the default value from `OFF` to `ON`.",
                source_pr=pr_url,
            )
        ],
        needs_review=False,
        reason="The PR changes a user-facing system variable.",
    )


def generated_note(note_type: str, release_note: str, pr_url: str):
    return models.GeneratedNote(
        note_type=note_type,
        release_note=release_note,
        needs_review=False,
        reason="Test result",
        doc_impact=detected_impact(pr_url),
    )


def append_source_header(sheet) -> None:
    sheet.append(
        [
            "component",
            "pr_author",
            "pr_link",
            "pr_title",
            "formated_release_note",
            "issue_type",
        ]
    )


def sample_row_context():
    pr_url = "https://github.com/pingcap/tidb/pull/456"
    issue_url = "https://github.com/pingcap/tidb/issues/123"
    return models.RowContext(
        row_number=2,
        component="TiDB",
        raw_component="tidb",
        issue_type="improvement",
        pr_title="Improve an operation",
        pr_authors=["contributor"],
        pr_urls=[pr_url],
        issue_urls=[issue_url],
        formatted_release_note="- Draft release note",
        issues=[
            models.IssueInfo(
                url=issue_url,
                title="Issue title",
                body="Issue body",
                labels=[],
            )
        ],
        pulls=[
            models.PullInfo(
                url=pr_url,
                title="PR title",
                body="PR body",
                author="contributor",
                head_ref="feature",
                base_ref="master",
                files_summary="file: config.go\npatch:\n+ tidb_example = ON",
            )
        ],
    )


class DocImpactValidationTest(unittest.TestCase):
    def test_validate_detected_impact_and_format_for_excel(self):
        issue_url = "https://github.com/pingcap/tidb/issues/123"
        pr_url = "https://github.com/pingcap/tidb/pull/456"
        data = {
            "type": "improvement",
            "release_note": (
                "- Improve an operation "
                f"[#123]({issue_url}) "
                "@[contributor](https://github.com/contributor)"
            ),
            "needs_review": False,
            "reason": "User-visible improvement.",
            "variable_or_config_doc_impact": {
                "status": "detected",
                "changes": [
                    {
                        "kind": "system_variable",
                        "name": "tidb_example",
                        "change_type": "modified",
                        "description": "Changes the default value from `OFF` to `ON`.",
                        "source_pr": pr_url,
                    }
                ],
                "needs_review": False,
                "reason": "The default value changes.",
            },
        }

        result, errors = ai_client.validate_ai_response(
            data,
            [issue_url],
            ["contributor"],
            [pr_url],
        )

        self.assertEqual([], errors)
        self.assertIsNotNone(result)
        formatted = excel_workbook.format_doc_impact_for_excel(result.doc_impact)
        self.assertIn("Detected | System variable | tidb_example | Modified", formatted)
        self.assertIn(pr_url, formatted)

    def test_rejects_source_pr_outside_the_input_row(self):
        impact = {
            "status": "detected",
            "changes": [
                {
                    "kind": "configuration_parameter",
                    "name": "raftstore.example",
                    "change_type": "newly_added",
                    "description": "Adds a user-facing setting.",
                    "source_pr": "https://github.com/tikv/tikv/pull/999",
                }
            ],
            "needs_review": False,
            "reason": "A parameter is added.",
        }

        result, errors = ai_client.validate_doc_impact(
            impact,
            ["https://github.com/tikv/tikv/pull/123"],
        )

        self.assertIsNone(result)
        self.assertTrue(any("source_pr" in error for error in errors))

    def test_uncertain_impact_requires_review(self):
        impact = {
            "status": "uncertain",
            "changes": [],
            "needs_review": False,
            "reason": "The diff is truncated.",
        }

        result, errors = ai_client.validate_doc_impact(impact, [])

        self.assertIsNone(result)
        self.assertTrue(any("must be true" in error for error in errors))

    def test_doc_only_analysis_does_not_revalidate_existing_release_note(self):
        pr_url = "https://github.com/pingcap/tidb/pull/456"
        data = {
            "type": "unexpected",
            "release_note": "This value is ignored for a doc-only task.",
            "needs_review": "invalid but ignored",
            "reason": None,
            "variable_or_config_doc_impact": {
                "status": "not_detected",
                "changes": [],
                "needs_review": False,
                "reason": "No user-facing variable or parameter changes.",
            },
        }

        result, errors = ai_client.validate_ai_response(
            data,
            [pr_url],
            ["contributor"],
            [pr_url],
            validate_release_note=False,
        )

        self.assertEqual([], errors)
        self.assertIsNotNone(result)
        self.assertEqual("not_detected", result.doc_impact.status)
        self.assertIsNone(result.note_type)
        self.assertIsNone(result.release_note)
        self.assertIsNone(result.needs_review)
        self.assertIsNone(result.reason)

    def test_release_note_only_analysis_does_not_revalidate_existing_doc_impact(self):
        pr_url = "https://github.com/pingcap/tidb/pull/456"
        data = {
            "type": "improvement",
            "release_note": (
                "- Improve an operation "
                f"[#456]({pr_url}) "
                "@[contributor](https://github.com/contributor)"
            ),
            "needs_review": False,
            "reason": "User-visible improvement.",
        }

        result, errors = ai_client.validate_ai_response(
            data,
            [pr_url],
            ["contributor"],
            [pr_url],
            validate_doc_impact_result=False,
        )

        self.assertEqual([], errors)
        self.assertIsNotNone(result)
        self.assertIsNone(result.doc_impact)

    def test_missing_doc_impact_after_validation_returns_an_error(self):
        pr_url = "https://github.com/pingcap/tidb/pull/456"
        data = {
            "variable_or_config_doc_impact": {
                "status": "not_detected",
                "changes": [],
                "needs_review": False,
                "reason": "No user-facing changes.",
            }
        }

        with mock.patch.object(
            ai_client,
            "validate_doc_impact",
            return_value=(None, []),
        ):
            result, errors = ai_client.validate_ai_response(
                data,
                [pr_url],
                [],
                [pr_url],
                validate_release_note=False,
            )

        self.assertIsNone(result)
        self.assertEqual(
            ["variable_or_config_doc_impact validation returned no result"],
            errors,
        )

    def test_detected_status_requires_a_valid_change(self):
        impact = {
            "status": "detected",
            "changes": ["not an object"],
            "needs_review": False,
            "reason": "A change was reported.",
        }

        result, errors = ai_client.validate_doc_impact(impact, [])

        self.assertIsNone(result)
        self.assertTrue(any("must be an object" in error for error in errors))
        self.assertTrue(
            any("at least one valid change" in error for error in errors)
        )

    def test_doc_only_prompt_and_schema_omit_release_note_output(self):
        context = sample_row_context()
        prompt = ai_client.build_generation_prompt(
            context,
            context.issue_urls,
            ["contributor"],
            generate_release_note=False,
            generate_doc_impact=True,
        )
        schema = ai_client.ai_output_schema(
            include_release_note=False,
            include_doc_impact=True,
        )

        self.assertIn("Variable or configuration documentation-impact task", prompt)
        self.assertNotIn("Release-note task", prompt)
        self.assertNotIn("Expected links to include", prompt)
        self.assertNotIn("formatted_release_note_from_excel", prompt)
        self.assertEqual(["variable_or_config_doc_impact"], schema["required"])
        self.assertEqual(
            {"variable_or_config_doc_impact"},
            set(schema["properties"]),
        )
        self.assertLess(
            len(prompt),
            len(
                ai_client.build_generation_prompt(
                    context,
                    context.issue_urls,
                    ["contributor"],
                )
            ),
        )

    def test_release_only_prompt_and_schema_omit_doc_impact_output(self):
        context = sample_row_context()
        prompt = ai_client.build_generation_prompt(
            context,
            context.issue_urls,
            ["contributor"],
            generate_release_note=True,
            generate_doc_impact=False,
        )
        schema = ai_client.ai_output_schema(
            include_release_note=True,
            include_doc_impact=False,
        )

        self.assertIn("Release-note task", prompt)
        self.assertNotIn("Variable or configuration documentation-impact task", prompt)
        self.assertIn("Expected links to include", prompt)
        self.assertIn("formatted_release_note_from_excel", prompt)
        self.assertIn(
            "use it as the primary basis for choosing between `bug_fix` and "
            "`improvement`",
            prompt,
        )
        self.assertIn(
            "When `issue_type_from_excel` is empty, determine the type from all "
            "available context",
            prompt,
        )
        self.assertIn(
            "Focus bug-fix entries on user-visible symptoms or error messages, "
            "trigger conditions, and impact, and avoid including internal "
            "implementation details unless they are necessary to explain the "
            "user-visible behavior",
            prompt,
        )
        self.assertNotIn("variable_or_config_doc_impact", schema["properties"])
        self.assertEqual(
            {"type", "release_note", "needs_review", "reason"},
            set(schema["required"]),
        )

    def test_doc_only_ai_call_accepts_a_minimal_response(self):
        pr_url = "https://github.com/pingcap/tidb/pull/456"
        client = RecordingAIClient(
            {
                "variable_or_config_doc_impact": {
                    "status": "not_detected",
                    "changes": [],
                    "needs_review": False,
                    "reason": "No user-facing variable or parameter changes.",
                }
            }
        )

        result = client.generate(
            "doc-only prompt",
            [pr_url],
            ["contributor"],
            [pr_url],
            generate_release_note=False,
            generate_doc_impact=True,
        )

        self.assertIsNone(result.release_note)
        self.assertEqual("not_detected", result.doc_impact.status)
        self.assertEqual(
            ["variable_or_config_doc_impact"],
            client.schemas[0]["required"],
        )

    def test_combined_prompt_and_schema_include_both_outputs(self):
        context = sample_row_context()
        prompt = ai_client.build_generation_prompt(
            context,
            context.issue_urls,
            ["contributor"],
        )
        schema = ai_client.ai_output_schema()

        self.assertIn("Release-note task", prompt)
        self.assertIn("Variable or configuration documentation-impact task", prompt)
        self.assertIn("Expected links to include", prompt)
        self.assertEqual(
            {
                "type",
                "release_note",
                "needs_review",
                "reason",
                "variable_or_config_doc_impact",
            },
            set(schema["required"]),
        )


class WorkbookDocImpactTest(unittest.TestCase):
    def test_main_and_not_needed_sheets_both_keep_doc_impact(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "pr_for_release_note"
        append_source_header(sheet)
        improvement_pr = "https://github.com/pingcap/tidb/pull/101"
        not_needed_pr = "https://github.com/pingcap/tidb/pull/102"
        duplicate_pr = "https://github.com/pingcap/tidb/pull/103"
        sheet.append(["tidb", "contributor", improvement_pr, "Improve", "", "improvement"])
        sheet.append(["tidb", "contributor", not_needed_pr, "Refactor", "", "improvement"])
        sheet.append(["tidb", "contributor", duplicate_pr, "Duplicate", "", "improvement"])
        header = excel_workbook.prepare_sheet_columns(sheet)
        self.assertEqual(
            header["ai_note_type"] + 1,
            header[excel_workbook.DOC_IMPACT_HEADER],
        )
        doc_impact_column = get_column_letter(
            header[excel_workbook.DOC_IMPACT_HEADER]
        )
        self.assertEqual(
            excel_workbook.DOC_IMPACT_COLUMN_WIDTH,
            sheet.column_dimensions[doc_impact_column].width,
        )
        sheet.cell(
            row=4,
            column=header["published_release_notes"],
            value=(
                "- (dup): release-8.5.0.md Improvements + TiDB "
                "- Improve existing behavior"
            ),
        )
        ai = FakeAIClient(
            [
                generated_note(
                    "improvement",
                    (
                        "- Improve an operation "
                        f"[#101]({improvement_pr}) "
                        "@[contributor](https://github.com/contributor)"
                    ),
                    improvement_pr,
                ),
                generated_note(
                    "not_needed",
                    "Release note is not needed: internal refactor",
                    not_needed_pr,
                ),
                generated_note(
                    "improvement",
                    (
                        "- Improve existing behavior "
                        f"[#103]({duplicate_pr}) "
                        "@[contributor](https://github.com/contributor)"
                    ),
                    duplicate_pr,
                ),
            ]
        )

        excel_workbook.generate_notes_for_sheet(
            sheet,
            header,
            FakeGitHubClient(),
            ai,
            ai_workers=1,
            github_workers=1,
        )

        self.assertEqual(3, ai.calls)
        self.assertTrue(
            sheet.cell(row=2, column=header[excel_workbook.DOC_IMPACT_HEADER])
            .value
            .startswith("Detected |")
        )
        self.assertTrue(
            sheet.cell(
                row=2, column=header[excel_workbook.DOC_IMPACT_HEADER]
            ).alignment.wrap_text
        )
        self.assertTrue(
            sheet.cell(row=3, column=header[excel_workbook.DOC_IMPACT_HEADER])
            .value
            .startswith("Detected |")
        )
        self.assertTrue(
            sheet.cell(row=4, column=header[excel_workbook.DOC_IMPACT_HEADER])
            .value
            .startswith("Detected |")
        )
        self.assertIsNone(
            sheet.cell(row=4, column=header["release_notes_written_by_ai"]).value
        )

        moved = excel_workbook.move_not_needed_rows_to_sheet(workbook, sheet, header)

        self.assertEqual(1, moved)
        target = workbook["release_note_not_needed"]
        target_header = excel_workbook.get_header(target)
        self.assertIn(excel_workbook.DOC_IMPACT_HEADER, target_header)
        target_doc_impact_column = get_column_letter(
            target_header[excel_workbook.DOC_IMPACT_HEADER]
        )
        self.assertEqual(
            excel_workbook.DOC_IMPACT_COLUMN_WIDTH,
            target.column_dimensions[target_doc_impact_column].width,
        )
        self.assertTrue(
            target.cell(
                row=2,
                column=target_header[excel_workbook.DOC_IMPACT_HEADER],
            )
            .value
            .startswith("Detected |")
        )

        doc_only_resume_ai = FakeAIClient([])
        excel_workbook.generate_notes_for_sheet(
            target,
            target_header,
            FakeGitHubClient(),
            doc_only_resume_ai,
            ai_workers=1,
            github_workers=1,
        )
        self.assertEqual(0, doc_only_resume_ai.calls)

    def test_existing_not_needed_sheet_gets_missing_impact_without_overwriting_note(self):
        workbook = Workbook()
        target = workbook.active
        target.title = "release_note_not_needed"
        append_source_header(target)
        target.cell(row=1, column=7, value="release_notes_written_by_ai")
        target.cell(row=1, column=8, value="ai_note_type")
        target.cell(row=1, column=9, value="published_release_notes")
        pr_url = "https://github.com/pingcap/tidb/pull/201"
        target.append(
            [
                "tidb",
                "contributor",
                pr_url,
                "Refactor",
                "",
                "improvement",
                "Release note is not needed: internal refactor",
                "not_needed",
                "",
            ]
        )
        header = excel_workbook.prepare_sheet_columns(target)
        original_note = target.cell(
            row=2, column=header["release_notes_written_by_ai"]
        ).value
        ai = FakeAIClient(
            [
                generated_note(
                    "not_needed",
                    "Release note is not needed: internal refactor",
                    pr_url,
                )
            ]
        )

        excel_workbook.generate_notes_for_sheet(
            target,
            header,
            FakeGitHubClient(),
            ai,
            ai_workers=1,
            github_workers=1,
        )

        self.assertEqual(1, ai.calls)
        self.assertEqual([(False, True)], ai.task_flags)
        self.assertNotIn("Release-note task", ai.prompts[0])
        self.assertIn(
            "Variable or configuration documentation-impact task",
            ai.prompts[0],
        )
        self.assertEqual(
            original_note,
            target.cell(row=2, column=header["release_notes_written_by_ai"]).value,
        )
        self.assertTrue(
            target.cell(row=2, column=header[excel_workbook.DOC_IMPACT_HEADER])
            .value
            .startswith("Detected |")
        )


if __name__ == "__main__":
    unittest.main()
