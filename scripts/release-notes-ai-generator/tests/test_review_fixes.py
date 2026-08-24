from __future__ import annotations

import argparse
import importlib
import os
import sys
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

from openpyxl import Workbook


SCRIPTS_DIR = Path(__file__).resolve().parents[2]
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

ai_client = importlib.import_module("release-notes-ai-generator.ai_client")
cli = importlib.import_module("release-notes-ai-generator.cli")
excel_workbook = importlib.import_module("release-notes-ai-generator.excel_workbook")
github_client = importlib.import_module("release-notes-ai-generator.github_client")
scope_filter = importlib.import_module("release-notes-ai-generator.scope_filter")


class AzureOpenAIClientTest(unittest.TestCase):
    def create_client(self, response):
        responses = SimpleNamespace(create=mock.Mock(return_value=response))
        sdk_client = SimpleNamespace(responses=responses)
        environment = {
            "AZURE_OPENAI_KEY": "test-key",
            "AZURE_OPENAI_BASE_URL": "https://example.openai.azure.com",
            "AZURE_OPENAI_DEPLOYMENT": "release-notes-deployment",
        }
        with (
            mock.patch.dict(os.environ, environment, clear=True),
            mock.patch("openai.OpenAI", return_value=sdk_client) as openai_mock,
        ):
            client = ai_client.AzureOpenAIClient(None, 30)
        return client, responses, openai_mock

    def test_uses_azure_v1_endpoint_and_structured_output(self):
        response = SimpleNamespace(
            status="completed",
            incomplete_details=None,
            error=None,
            output_text='{"type":"improvement"}',
        )
        client, responses, openai_mock = self.create_client(response)
        schema = {"type": "object", "properties": {}, "required": []}

        output = client._run("test prompt", schema)

        self.assertEqual('{"type":"improvement"}', output)
        self.assertEqual(
            "https://example.openai.azure.com/openai/v1/",
            openai_mock.call_args.kwargs["base_url"],
        )
        request = responses.create.call_args.kwargs
        self.assertEqual(schema, request["text"]["format"]["schema"])
        self.assertTrue(request["text"]["format"]["strict"])

    def test_rejects_incomplete_response_with_reason(self):
        response = SimpleNamespace(
            status="incomplete",
            incomplete_details=SimpleNamespace(reason="max_output_tokens"),
            error=None,
            output_text="",
        )
        client, _responses, _openai_mock = self.create_client(response)

        with self.assertRaisesRegex(RuntimeError, "max_output_tokens"):
            client._run("test prompt", {"type": "object"})

    def test_requires_an_explicit_azure_deployment(self):
        environment = {
            "AZURE_OPENAI_KEY": "test-key",
            "AZURE_OPENAI_BASE_URL": "https://example.openai.azure.com",
        }
        with mock.patch.dict(os.environ, environment, clear=True):
            with self.assertRaisesRegex(ValueError, "deployment name"):
                ai_client.AzureOpenAIClient(None, 30)


class CliValidationTest(unittest.TestCase):
    def test_rejects_non_positive_ai_timeout_before_loading_github_token(self):
        args = argparse.Namespace(
            ai_workers=1,
            github_workers=1,
            author_workers=1,
            checkpoint_interval=1,
            scope_base_branch_start_date=None,
            involve_ai_generation="ON",
            ai_timeout=0,
        )
        with mock.patch.object(cli, "load_github_token") as load_token:
            with self.assertRaisesRegex(ValueError, "--ai-timeout"):
                cli.run_generate(args)
        load_token.assert_not_called()


class WorkbookCopyTest(unittest.TestCase):
    def test_sorting_updates_hyperlink_reference(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["component", "link"])
        sheet.append(["z", "Z link"])
        sheet.append(["a", "A link"])
        sheet["B2"].hyperlink = "https://example.com/z"

        excel_workbook.sort_sheet_rows_by_component(sheet)

        self.assertEqual("z", sheet["A3"].value)
        self.assertEqual("B3", sheet["B3"].hyperlink.ref)

    def test_scope_archive_maps_reordered_columns_by_header(self):
        workbook = Workbook()
        source = workbook.active
        source.title = "source"
        source.append(["a", "b", "c"])
        source.append(["value-a", "value-b", "value-c"])
        target = workbook.create_sheet("target")
        target.append(["b", "Reason", "a"])

        scope_filter.append_row_with_reason(source, target, 2, "out of scope")

        header = excel_workbook.get_header(target)
        self.assertEqual("value-a", target.cell(2, header["a"]).value)
        self.assertEqual("value-b", target.cell(2, header["b"]).value)
        self.assertEqual("value-c", target.cell(2, header["c"]).value)
        self.assertEqual("out of scope", target.cell(2, header["Reason"]).value)

    def test_same_series_archive_maps_reordered_columns_by_header(self):
        workbook = Workbook()
        source = workbook.active
        source.title = "source"
        source.append(["a", "b", "c"])
        source.append(["value-a", "value-b", "value-c"])
        target = workbook.create_sheet("target")
        target.append(["b", "reason", "a"])
        reason_col = excel_workbook.ensure_same_series_reason_header(
            source,
            target,
        )

        excel_workbook.append_row_with_reason(
            source,
            target,
            2,
            "already published",
            reason_col,
        )

        header = excel_workbook.get_header(target)
        self.assertEqual("value-a", target.cell(2, header["a"]).value)
        self.assertEqual("value-b", target.cell(2, header["b"]).value)
        self.assertEqual("value-c", target.cell(2, header["c"]).value)
        self.assertEqual("already published", target.cell(2, header["reason"]).value)


class GitHubClientTest(unittest.TestCase):
    def setUp(self):
        self.client = github_client.GitHubClient(None)

    def test_get_pull_can_skip_changed_files(self):
        pull_data = {
            "title": "Test pull",
            "body": "Body",
            "user": {"login": "author"},
            "head": {"ref": "feature"},
            "base": {"ref": "master"},
        }
        with (
            mock.patch.object(self.client, "get_json", return_value=pull_data),
            mock.patch.object(self.client, "get_pull_files_summary") as files,
        ):
            pull = self.client.get_pull(
                "https://github.com/pingcap/tidb/pull/1",
                include_files=False,
            )

        self.assertEqual("", pull.files_summary)
        files.assert_not_called()

    def test_changed_file_pagination_stops_after_a_short_page(self):
        page = [
            {
                "filename": "test.py",
                "status": "modified",
                "additions": 1,
                "deletions": 0,
                "patch": "+change",
            }
        ]
        with mock.patch.object(
            self.client,
            "get_api_json",
            return_value=page,
        ) as get_api_json:
            summary = self.client.get_pull_files_summary("owner", "repo", "1")

        self.assertIn("file: test.py", summary)
        get_api_json.assert_called_once()

    def test_pull_listing_reports_direction_and_completeness(self):
        page = [
            {
                "html_url": "https://github.com/pingcap/tidb/pull/1",
                "title": "Test pull",
                "user": {"login": "author"},
                "head": {"ref": "feature"},
                "base": {"ref": "release-8.5"},
            }
        ]
        with mock.patch.object(
            self.client,
            "get_api_json",
            return_value=page,
        ) as get_api_json:
            pulls, truncated = self.client.list_pulls_for_base_with_state(
                "pingcap",
                "tidb",
                "release-8.5",
                direction="desc",
            )

        self.assertEqual(1, len(pulls))
        self.assertFalse(truncated)
        self.assertEqual("desc", get_api_json.call_args.kwargs["params"]["direction"])


if __name__ == "__main__":
    unittest.main()
