from __future__ import annotations

import copy
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import OrderedDict
from pathlib import Path
from typing import Any, Callable

from openpyxl.styles import PatternFill

from .ai_client import build_generation_prompt
from .constants import (
    AUTHOR_RE,
    BOT_AUTHORS,
    COMPONENT_HEADERS,
    GITHUB_ITEM_URL_RE,
    REQUIRED_HEADERS,
    TOOL_COMPONENTS,
    TOP_LEVEL_COMPONENTS,
)
from .models import (
    ExistingNote,
    GitHubDataCache,
    MarkdownEntry,
    RowContext,
    RowGenerationResult,
    RowInput,
)
from .utils import (
    extract_issue_urls,
    extract_pr_urls,
    normalize_component,
    normalize_raw_component,
    normalized_release_component,
    replace_author_markdown,
    split_lines,
    split_multi_value,
    str_value,
    unique_ordered,
)


GRAY_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
NOT_NEEDED_PREFIX = "Release note is not needed:"
SAME_SERIES_REASON_HEADER = "reason"


def prepare_sheet_columns(sheet: Any) -> dict[str, int]:
    header = get_header(sheet)
    missing = sorted(REQUIRED_HEADERS - set(header))
    if missing:
        raise ValueError(f"Missing required Excel columns: {', '.join(missing)}")
    get_component_col(header)

    ai_col = header.get("release_notes_written_by_ai")
    formatted_col = header["formated_release_note"]
    if not ai_col:
        sheet.insert_cols(formatted_col + 1)
        sheet.cell(row=1, column=formatted_col + 1, value="release_notes_written_by_ai")
        header = get_header(sheet)

    if "published_release_notes" not in header:
        last_col = sheet.max_column
        sheet.cell(row=1, column=last_col + 1, value="published_release_notes")
        header = get_header(sheet)
    return header


def get_header(sheet: Any) -> dict[str, int]:
    header: dict[str, int] = {}
    for index, cell in enumerate(sheet[1], start=1):
        if cell.value:
            header[str(cell.value).strip()] = index
    return header


def clear_output_columns(sheet: Any, header: dict[str, int], clear_ai: bool = True) -> None:
    for row_number in range(2, sheet.max_row + 1):
        if clear_ai:
            sheet.cell(row=row_number, column=header["release_notes_written_by_ai"]).value = None
        sheet.cell(row=row_number, column=header["published_release_notes"]).value = None


def sort_sheet_rows_by_component(sheet: Any) -> None:
    header = get_header(sheet)
    component_col = get_component_col(header)
    if sheet.max_row <= 2:
        return

    snapshots = [
        (row_number, component_sort_key(sheet.cell(row=row_number, column=component_col).value), snapshot_row(sheet, row_number))
        for row_number in range(2, sheet.max_row + 1)
    ]
    sorted_snapshots = sorted(snapshots, key=lambda item: item[1])
    if [row_number for row_number, _key, _snapshot in snapshots] == [
        row_number for row_number, _key, _snapshot in sorted_snapshots
    ]:
        return

    for target_row, (_source_row, _key, snapshot) in enumerate(sorted_snapshots, start=2):
        restore_row(sheet, target_row, snapshot)

    print("Sorted worksheet rows by component before release-note generation", flush=True)


def component_sort_key(value: Any) -> tuple[int, str]:
    component = normalize_raw_component(value)
    if not component:
        return (1, "")
    return (0, component.casefold())


def snapshot_row(sheet: Any, row_number: int) -> dict[str, Any]:
    row_dimension = sheet.row_dimensions[row_number]
    return {
        "height": row_dimension.height,
        "hidden": row_dimension.hidden,
        "outline_level": row_dimension.outlineLevel,
        "collapsed": row_dimension.collapsed,
        "cells": [snapshot_cell(sheet.cell(row=row_number, column=column)) for column in range(1, sheet.max_column + 1)],
    }


def snapshot_cell(cell: Any) -> dict[str, Any]:
    return {
        "value": cell.value,
        "style": copy.copy(cell._style),
        "number_format": cell.number_format,
        "hyperlink": copy.copy(cell.hyperlink) if cell.hyperlink else None,
        "comment": copy.copy(cell.comment) if cell.comment else None,
    }


def restore_row(sheet: Any, row_number: int, snapshot: dict[str, Any]) -> None:
    row_dimension = sheet.row_dimensions[row_number]
    row_dimension.height = snapshot["height"]
    row_dimension.hidden = snapshot["hidden"]
    row_dimension.outlineLevel = snapshot["outline_level"]
    row_dimension.collapsed = snapshot["collapsed"]
    for column, cell_snapshot in enumerate(snapshot["cells"], start=1):
        cell = sheet.cell(row=row_number, column=column)
        cell.value = cell_snapshot["value"]
        cell._style = copy.copy(cell_snapshot["style"])
        cell.number_format = cell_snapshot["number_format"]
        cell._hyperlink = copy.copy(cell_snapshot["hyperlink"]) if cell_snapshot["hyperlink"] else None
        cell.comment = copy.copy(cell_snapshot["comment"]) if cell_snapshot["comment"] else None


def get_component_col(header: dict[str, int]) -> int:
    for name in COMPONENT_HEADERS:
        if name in header:
            return header[name]
    raise ValueError("Missing required Excel column: component or components")


def issue_urls_for_row(sheet: Any, header: dict[str, int], row_number: int) -> list[str]:
    candidates: list[str] = []
    if "issue_url" in header:
        candidates.append(str_value(sheet.cell(row=row_number, column=header["issue_url"]).value))
    candidates.append(str_value(sheet.cell(row=row_number, column=header["formated_release_note"]).value))
    return unique_ordered(url for text in candidates for url in extract_issue_urls(text))


def first_issue_url_for_row(sheet: Any, header: dict[str, int], row_number: int) -> str | None:
    issue_urls = issue_urls_for_row(sheet, header, row_number)
    return issue_urls[0] if issue_urls else None


def store_existing_release_notes(releases_dir: Path, version: str) -> list[ExistingNote]:
    existing_notes: list[ExistingNote] = []
    seen: set[tuple[str, tuple[str, ...], str]] = set()
    target_version = parse_semver_tuple(version)

    for file_path in sorted(releases_dir.rglob("*.md")):
        if should_skip_release_file(file_path, target_version):
            continue
        level1 = level2 = level3 = ""
        with file_path.open("r", encoding="utf-8") as file:
            for raw_line in file:
                line = raw_line.strip()
                authors = AUTHOR_RE.findall(line)
                item_urls = [match.group() for match in GITHUB_ITEM_URL_RE.finditer(line)]
                if item_urls:
                    note_level = level1 + level2 + level3
                    note_type, component = classify_note_level(note_level)
                    for item_url in item_urls:
                        key = (item_url, tuple(authors), file_path.name)
                        if key in seen:
                            continue
                        seen.add(key)
                        existing_notes.append(
                            ExistingNote(
                                url=item_url,
                                line=line,
                                file_name=file_path.name,
                                note_level=note_level,
                                authors=authors,
                                note_type=note_type,
                                component=component,
                            )
                        )
                    continue

                heading = parse_release_note_heading(raw_line)
                if not heading:
                    continue
                heading_level, label = heading
                if heading_level == 1:
                    level1 = "> " + label
                    level2 = level3 = ""
                elif heading_level == 2:
                    level2 = "> " + label
                    level3 = ""
                elif heading_level == 3:
                    level3 = "> " + label
    return existing_notes


def should_skip_release_file(file_path: Path, target_version: tuple[int, int, int]) -> bool:
    if "updated-by-ai" in file_path.stem:
        return True
    file_version = release_file_semver_tuple(file_path)
    if not file_version:
        return False
    return file_version >= target_version


def parse_semver_tuple(version: str) -> tuple[int, int, int]:
    match = re.match(r"^(?P<major>\d+)\.(?P<minor>\d+)\.(?P<patch>\d+)", version)
    if not match:
        raise ValueError(f"Invalid TiDB version: {version}")
    return (
        int(match.group("major")),
        int(match.group("minor")),
        int(match.group("patch")),
    )


def release_file_semver_tuple(file_path: Path) -> tuple[int, int, int] | None:
    match = re.match(
        r"^release-(?P<major>\d+)\.(?P<minor>\d+)\.(?P<patch>\d+)",
        file_path.stem,
    )
    if not match:
        return None
    return (
        int(match.group("major")),
        int(match.group("minor")),
        int(match.group("patch")),
    )


def parse_release_note_heading(raw_line: str) -> tuple[int, str] | None:
    line = raw_line.rstrip()
    section = re.match(r"^##\s+(.+?)\s*$", line)
    if section:
        return 1, section.group(1).strip()

    top_component = re.match(r"^[+-]\s+(.+?)\s*$", line)
    if top_component:
        label = top_component.group(1).strip()
        if label.lower() == "tools" or normalized_release_component(label):
            return 2, label

    tool_component = re.match(r"^ {4}[+-]\s+(.+?)\s*$", line)
    if tool_component:
        label = tool_component.group(1).strip()
        if normalized_release_component(label):
            return 3, label
    return None


def update_pr_authors_and_dup_notes(
    sheet: Any,
    header: dict[str, int],
    existing_notes: list[ExistingNote],
    github: Any,
    author_workers: int = 1,
) -> None:
    apply_bot_author_replacements(sheet, header, github, author_workers)
    existing_notes_by_url = index_existing_notes_by_url(existing_notes)

    for row_number in range(2, sheet.max_row + 1):
        author_cell = sheet.cell(row=row_number, column=header["pr_author"])
        current_author = str_value(author_cell.value)

        issue_urls = issue_urls_for_row(sheet, header, row_number)
        if not issue_urls:
            continue

        current_authors = split_multi_value(current_author)
        dup_notes = []
        for issue_url in issue_urls:
            for existing in existing_notes_by_url.get(issue_url, []):
                if existing.authors and not set(current_authors).intersection(existing.authors):
                    continue
                dup_notes.append(existing.dup_text)

        if dup_notes:
            dup_col = header["published_release_notes"]
            sheet.cell(row=row_number, column=dup_col, value="\n".join(unique_ordered(dup_notes)))
            fill_row(sheet, row_number)
            print(
                f"Row {row_number}: found duplicated release note for {', '.join(issue_urls)}",
                flush=True,
            )


def move_rows_with_issues_already_in_same_series(
    workbook: Any,
    sheet: Any,
    header: dict[str, int],
    existing_notes: list[ExistingNote],
    version: str,
) -> int:
    files_by_issue_url = same_series_release_files_by_issue_url(existing_notes, version)
    if not files_by_issue_url:
        return 0

    target_sheet_name = same_series_issues_sheet_name(version)
    target, reason_col = ensure_sheet_with_reason(workbook, sheet, target_sheet_name)
    rows_to_move: list[tuple[int, str]] = []

    for row_number in range(2, sheet.max_row + 1):
        issue_urls = issue_urls_for_row(sheet, header, row_number)
        reason = same_series_issue_reason(issue_urls, files_by_issue_url)
        if reason:
            rows_to_move.append((row_number, reason))

    for row_number, reason in rows_to_move:
        append_row_with_reason(sheet, target, row_number, reason, reason_col)

    for row_number, _reason in reversed(rows_to_move):
        sheet.delete_rows(row_number, 1)

    if rows_to_move:
        print(
            f"Moved {len(rows_to_move)} row(s) to {target_sheet_name} because their issues "
            "already appear in earlier release notes from the same major.minor series",
            flush=True,
        )
    return len(rows_to_move)


def move_not_needed_rows_to_sheet(
    workbook: Any,
    sheet: Any,
    header: dict[str, int],
) -> int:
    """Move rows where AI determined no release note is needed to a separate sheet."""
    ai_col = header["release_notes_written_by_ai"]
    target_sheet_name = "release_note_not_needed"

    rows_to_move: list[int] = []
    for row_number in range(2, sheet.max_row + 1):
        ai_value = str_value(sheet.cell(row=row_number, column=ai_col).value)
        if ai_value.startswith(NOT_NEEDED_PREFIX):
            rows_to_move.append(row_number)

    if not rows_to_move:
        return 0

    if target_sheet_name in workbook.sheetnames:
        target = workbook[target_sheet_name]
        if not str_value(target.cell(row=1, column=1).value):
            copy_header_row(sheet, target)
    else:
        target = workbook.create_sheet(target_sheet_name)
        copy_header_row(sheet, target)

    for row_number in rows_to_move:
        target_row = target.max_row + 1
        for column in range(1, sheet.max_column + 1):
            copy_cell(
                sheet.cell(row=row_number, column=column),
                target.cell(row=target_row, column=column),
            )

    for row_number in reversed(rows_to_move):
        sheet.delete_rows(row_number, 1)

    print(
        f"Moved {len(rows_to_move)} row(s) to sheet '{target_sheet_name}' "
        "(release note not needed)",
        flush=True,
    )
    return len(rows_to_move)


def copy_header_row(source_sheet: Any, target_sheet: Any) -> None:
    for column in range(1, source_sheet.max_column + 1):
        copy_cell(
            source_sheet.cell(row=1, column=column),
            target_sheet.cell(row=1, column=column),
        )


def same_series_release_files_by_issue_url(
    existing_notes: list[ExistingNote],
    version: str,
) -> dict[str, list[str]]:
    target_version = parse_semver_tuple(version)
    files_by_issue_url: dict[str, list[str]] = {}

    for existing in existing_notes:
        match = GITHUB_ITEM_URL_RE.search(existing.url)
        if not match or match.group("kind") != "issues":
            continue

        file_version = release_file_semver_tuple(Path(existing.file_name))
        if not file_version:
            continue
        if file_version[:2] != target_version[:2] or file_version >= target_version:
            continue

        files = files_by_issue_url.setdefault(existing.url, [])
        if existing.file_name not in files:
            files.append(existing.file_name)

    for issue_url, files in list(files_by_issue_url.items()):
        files_by_issue_url[issue_url] = sorted(files, key=release_file_name_sort_key)
    return files_by_issue_url


def same_series_issues_sheet_name(version: str) -> str:
    major, minor, _patch = parse_semver_tuple(version)
    return f"issue_already_in_earlier_v{major}.{minor}"


def same_series_issue_reason(
    issue_urls: list[str],
    files_by_issue_url: dict[str, list[str]],
) -> str | None:
    reasons = []
    for issue_url in issue_urls:
        files = files_by_issue_url.get(issue_url)
        if files:
            reasons.append(f"{issue_url} appears in {', '.join(files)}")
    return "; ".join(reasons) if reasons else None


def release_file_name_sort_key(file_name: str) -> tuple[int, int, int, str]:
    version = release_file_semver_tuple(Path(file_name))
    if not version:
        return (sys.maxsize, sys.maxsize, sys.maxsize, file_name)
    return (*version, file_name)


def ensure_sheet_with_reason(
    workbook: Any,
    source_sheet: Any,
    target_sheet_name: str,
) -> tuple[Any, int]:
    if target_sheet_name in workbook.sheetnames:
        target = workbook[target_sheet_name]
        if not str_value(target.cell(row=1, column=1).value):
            reason_col = copy_header_with_reason(source_sheet, target)
        else:
            reason_col = ensure_same_series_reason_header(source_sheet, target)
        return target, reason_col

    target = workbook.create_sheet(target_sheet_name)
    reason_col = copy_header_with_reason(source_sheet, target)
    return target, reason_col


def copy_header_with_reason(source_sheet: Any, target_sheet: Any) -> int:
    for column in range(1, source_sheet.max_column + 1):
        copy_cell(
            source_sheet.cell(row=1, column=column),
            target_sheet.cell(row=1, column=column),
        )
    return ensure_same_series_reason_header(source_sheet, target_sheet)


def ensure_same_series_reason_header(source_sheet: Any, target_sheet: Any) -> int:
    reason_col = find_header_column(target_sheet, SAME_SERIES_REASON_HEADER)
    if not reason_col:
        reason_col = max(source_sheet.max_column, target_sheet.max_column) + 1
        copy_missing_header_cells(source_sheet, target_sheet)
        target_sheet.cell(row=1, column=reason_col, value=SAME_SERIES_REASON_HEADER)
        return reason_col

    while reason_col <= source_sheet.max_column:
        target_sheet.insert_cols(reason_col)
        reason_col += 1

    copy_missing_header_cells(source_sheet, target_sheet)
    return reason_col


def copy_missing_header_cells(source_sheet: Any, target_sheet: Any) -> None:
    for column in range(1, source_sheet.max_column + 1):
        if not str_value(target_sheet.cell(row=1, column=column).value):
            copy_cell(
                source_sheet.cell(row=1, column=column),
                target_sheet.cell(row=1, column=column),
            )


def find_header_column(sheet: Any, header_name: str) -> int | None:
    for column in range(1, sheet.max_column + 1):
        if str_value(sheet.cell(row=1, column=column).value) == header_name:
            return column
    return None


def append_row_with_reason(
    source_sheet: Any,
    target_sheet: Any,
    row_number: int,
    reason: str,
    reason_col: int,
) -> None:
    target_row = target_sheet.max_row + 1
    source_dimension = source_sheet.row_dimensions[row_number]
    target_dimension = target_sheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.collapsed = source_dimension.collapsed

    for column in range(1, source_sheet.max_column + 1):
        copy_cell(
            source_sheet.cell(row=row_number, column=column),
            target_sheet.cell(row=target_row, column=column),
        )
    target_sheet.cell(row=target_row, column=reason_col, value=reason)


def copy_cell(source_cell: Any, target_cell: Any) -> None:
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell._style = copy.copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.hyperlink:
        target_cell._hyperlink = copy.copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy.copy(source_cell.comment)


def apply_bot_author_replacements(
    sheet: Any,
    header: dict[str, int],
    github: Any,
    author_workers: int,
) -> None:
    requests = bot_author_requests(sheet, header)
    if not requests:
        return
    print(
        f"Resolving {len(requests)} bot-authored PR row(s) with {author_workers} worker(s)",
        flush=True,
    )

    replacements = resolve_bot_author_replacements(requests, github, author_workers)
    for row_number in sorted(replacements):
        current_author, actual_author = replacements[row_number]
        author_cell = sheet.cell(row=row_number, column=header["pr_author"])
        formatted_cell = sheet.cell(row=row_number, column=header["formated_release_note"])
        formatted_note = str_value(formatted_cell.value)
        print(
            f"Replacing bot author in row {row_number}: {current_author} -> {actual_author}",
            flush=True,
        )
        author_cell.value = actual_author
        formatted_cell.value = replace_author_markdown(
            formatted_note, current_author, actual_author
        )


def bot_author_requests(sheet: Any, header: dict[str, int]) -> list[tuple[int, str, str, str]]:
    requests = []
    for row_number in range(2, sheet.max_row + 1):
        current_author = str_value(sheet.cell(row=row_number, column=header["pr_author"]).value)
        pr_link = str_value(sheet.cell(row=row_number, column=header["pr_link"]).value)
        if current_author not in BOT_AUTHORS or not pr_link:
            continue
        pr_title = str_value(sheet.cell(row=row_number, column=header["pr_title"]).value)
        requests.append((row_number, pr_link, pr_title, current_author))
    return requests


def resolve_bot_author_replacements(
    requests: list[tuple[int, str, str, str]],
    github: Any,
    author_workers: int,
) -> dict[int, tuple[str, str]]:
    replacements: dict[int, tuple[str, str]] = {}
    total = len(requests)
    if author_workers == 1:
        for completed, request in enumerate(requests, start=1):
            row_number, pr_link, pr_title, current_author = request
            actual_author = resolve_bot_author(github, request)
            print_bot_author_progress(completed, total, row_number, current_author, actual_author)
            if actual_author != current_author:
                replacements[row_number] = (current_author, actual_author)
        return replacements

    with ThreadPoolExecutor(max_workers=author_workers) as executor:
        futures = {
            executor.submit(resolve_bot_author, github, request): request
            for request in requests
        }
        for completed, future in enumerate(as_completed(futures), start=1):
            row_number, _pr_link, _pr_title, current_author = futures[future]
            actual_author = future.result()
            print_bot_author_progress(completed, total, row_number, current_author, actual_author)
            if actual_author != current_author:
                replacements[row_number] = (current_author, actual_author)
    return replacements


def print_bot_author_progress(
    completed: int,
    total: int,
    row_number: int,
    current_author: str,
    actual_author: str,
) -> None:
    status = "unchanged" if actual_author == current_author else f"{current_author} -> {actual_author}"
    print(
        f"Resolved bot author {completed}/{total}: row {row_number} ({status})",
        flush=True,
    )


def resolve_bot_author(github: Any, request: tuple[int, str, str, str]) -> str:
    row_number, pr_link, pr_title, current_author = request
    try:
        return github.get_original_author_for_cherry_pick(
            row_number,
            pr_link,
            pr_title,
            current_author,
        )
    except Exception as exc:  # noqa: BLE001
        print(
            f"Row {row_number}: failed to resolve bot author for {pr_link}: {exc}",
            file=sys.stderr,
            flush=True,
        )
        return current_author


def index_existing_notes_by_url(existing_notes: list[ExistingNote]) -> dict[str, list[ExistingNote]]:
    indexed: dict[str, list[ExistingNote]] = {}
    seen: set[tuple[str, tuple[str, ...]]] = set()
    for existing in existing_notes:
        key = (existing.url, tuple(existing.authors))
        if key in seen:
            continue
        seen.add(key)
        indexed.setdefault(existing.url, []).append(existing)
    return indexed


def merge_rows_by_issue_and_component(sheet: Any, header: dict[str, int]) -> None:
    groups: OrderedDict[tuple[str, str], list[int]] = OrderedDict()
    component_col = get_component_col(header)
    for row_number in range(2, sheet.max_row + 1):
        issue_url = first_issue_url_for_row(sheet, header, row_number)
        if not issue_url:
            continue
        component = normalize_raw_component(sheet.cell(row=row_number, column=component_col).value)
        if not component:
            continue
        groups.setdefault((issue_url, component), []).append(row_number)

    rows_to_delete: list[int] = []
    for (_issue_url, _component), rows in groups.items():
        if len(rows) <= 1:
            continue
        keep_row = rows[0]
        merge_pr_links(sheet, header, keep_row, rows)
        merge_authors(sheet, header, keep_row, rows)
        merge_dup_notes(sheet, header, keep_row, rows)
        fill_first_empty_values(sheet, header, keep_row, rows)
        if str_value(sheet.cell(row=keep_row, column=header["published_release_notes"]).value):
            fill_row(sheet, keep_row)
        rows_to_delete.extend(rows[1:])

    for row_number in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_number, 1)


def merge_pr_links(sheet: Any, header: dict[str, int], keep_row: int, rows: list[int]) -> None:
    links: list[str] = []
    for row in rows:
        links.extend(split_multi_value(sheet.cell(row=row, column=header["pr_link"]).value))
    sheet.cell(row=keep_row, column=header["pr_link"], value=", ".join(unique_ordered(links)))


def merge_authors(sheet: Any, header: dict[str, int], keep_row: int, rows: list[int]) -> None:
    authors: list[str] = []
    for row in rows:
        authors.extend(split_multi_value(sheet.cell(row=row, column=header["pr_author"]).value))
    sheet.cell(row=keep_row, column=header["pr_author"], value=", ".join(unique_ordered(authors)))


def merge_dup_notes(sheet: Any, header: dict[str, int], keep_row: int, rows: list[int]) -> None:
    notes: list[str] = []
    for row in rows:
        notes.extend(split_lines(sheet.cell(row=row, column=header["published_release_notes"]).value))
    if notes:
        sheet.cell(row=keep_row, column=header["published_release_notes"], value="\n".join(unique_ordered(notes)))


def fill_first_empty_values(sheet: Any, header: dict[str, int], keep_row: int, rows: list[int]) -> None:
    columns_to_skip = {
        header["pr_link"],
        header["pr_author"],
        header["published_release_notes"],
        header["release_notes_written_by_ai"],
    }
    for col in range(1, sheet.max_column + 1):
        if col in columns_to_skip:
            continue
        keep_cell = sheet.cell(row=keep_row, column=col)
        if str_value(keep_cell.value):
            continue
        for row in rows[1:]:
            value = sheet.cell(row=row, column=col).value
            if str_value(value):
                keep_cell.value = value
                break


def generate_notes_for_sheet(
    sheet: Any,
    header: dict[str, int],
    github: Any,
    ai: Any,
    ai_workers: int = 1,
    github_workers: int = 1,
    checkpoint_callback: Callable[[int, int], None] | None = None,
) -> list[MarkdownEntry]:
    entries_by_row: dict[int, list[MarkdownEntry]] = {}
    row_inputs = [
        build_row_input(sheet, header, row_number)
        for row_number in range(2, sheet.max_row + 1)
    ]
    rows_to_generate: list[RowInput] = []

    for row_input in row_inputs:
        row_number = row_input.row_number
        component = row_input.component
        dup_text = str_value(sheet.cell(row=row_number, column=header["published_release_notes"]).value)
        if dup_text:
            sheet.cell(row=row_number, column=header["release_notes_written_by_ai"]).value = None
            entries_by_row[row_number] = dup_entries_for_row(row_input, dup_text)
            continue

        ai_cell = sheet.cell(row=row_number, column=header["release_notes_written_by_ai"])
        expected_links = row_input.issue_urls or row_input.pr_urls
        if not expected_links:
            ai_cell.value = "AI_GENERATION_FAILED: missing issue URL and PR URL"
            continue

        existing_note = str_value(ai_cell.value)
        if is_reusable_ai_note(existing_note):
            if is_not_needed_note(existing_note):
                print(f"Row {row_number}: skipped existing not-needed verdict", flush=True)
                continue
            note_type = classify_note_type_from_text(existing_note, row_input.issue_type)
            entries_by_row[row_number] = [
                MarkdownEntry(
                    note_type or "improvement",
                    component,
                    existing_note,
                    row_input.raw_component,
                )
            ]
            print(f"Row {row_number}: skipped existing AI release note", flush=True)
            continue

        rows_to_generate.append(row_input)

    github_cache = prefetch_github_data(rows_to_generate, github, github_workers)
    total_to_generate = len(rows_to_generate)
    if total_to_generate:
        print(
            f"Generating AI release notes for {total_to_generate} row(s) "
            f"with {ai_workers} worker(s)",
            flush=True,
        )

    completed = 0
    with ThreadPoolExecutor(max_workers=ai_workers) as executor:
        futures = [
            executor.submit(generate_note_for_row, row_input, github_cache, ai)
            for row_input in rows_to_generate
        ]
        for future in as_completed(futures):
            result = future.result()
            apply_generation_result(sheet, header, result, entries_by_row)
            completed += 1
            if checkpoint_callback:
                checkpoint_callback(completed, total_to_generate)

    entries: list[MarkdownEntry] = []
    for row_input in row_inputs:
        entries.extend(entries_by_row.get(row_input.row_number, []))
    return entries


def generate_notes_without_ai(sheet: Any, header: dict[str, int]) -> list[MarkdownEntry]:
    entries: list[MarkdownEntry] = []
    for row_number in range(2, sheet.max_row + 1):
        row_input = build_row_input(sheet, header, row_number)
        dup_text = str_value(sheet.cell(row=row_number, column=header["published_release_notes"]).value)
        if dup_text:
            entries.extend(dup_entries_for_row(row_input, dup_text))
            continue

        formatted_notes = split_lines(row_input.formatted_release_note)
        if not formatted_notes:
            print(
                f"Row {row_number}: skipped non-dup row because formated_release_note is empty",
                file=sys.stderr,
                flush=True,
            )
            continue
        note_type = classify_note_type_from_text(
            row_input.formatted_release_note,
            row_input.issue_type,
        )
        for note in formatted_notes:
            entries.append(
                MarkdownEntry(
                    note_type or "improvement",
                    row_input.component,
                    note,
                    row_input.raw_component,
                )
            )

    print(
        f"AI generation is OFF; generated Markdown from formated_release_note for {len(entries)} note(s)",
        flush=True,
    )
    return entries


def dup_entries_for_row(row_input: RowInput, dup_text: str) -> list[MarkdownEntry]:
    entries: list[MarkdownEntry] = []
    for dup_note in split_lines(dup_text):
        note_type = classify_note_type_from_text(
            dup_note,
            row_input.issue_type,
        )
        dup_component = parse_component_from_dup(dup_note) or row_input.component
        if note_type in {"improvement", "bug_fix"}:
            entries.append(
                MarkdownEntry(
                    note_type,
                    normalize_component(dup_component),
                    dup_note,
                    row_input.raw_component,
                )
            )
    return entries


def build_row_input(sheet: Any, header: dict[str, int], row_number: int) -> RowInput:
    raw_component = normalize_raw_component(
        sheet.cell(row=row_number, column=get_component_col(header)).value
    )
    return RowInput(
        row_number=row_number,
        component=release_component_for_row(sheet, header, row_number),
        raw_component=raw_component,
        issue_type=str_value(sheet.cell(row=row_number, column=header["issue_type"]).value),
        pr_title=str_value(sheet.cell(row=row_number, column=header["pr_title"]).value),
        pr_authors=split_multi_value(sheet.cell(row=row_number, column=header["pr_author"]).value),
        pr_urls=extract_pr_urls(str_value(sheet.cell(row=row_number, column=header["pr_link"]).value)),
        issue_urls=issue_urls_for_row(sheet, header, row_number),
        formatted_release_note=str_value(
            sheet.cell(row=row_number, column=header["formated_release_note"]).value
        ),
    )


def is_reusable_ai_note(note: str) -> bool:
    return bool(note) and not note.startswith("AI_GENERATION_FAILED:")


def is_not_needed_note(note: str) -> bool:
    return note.startswith(NOT_NEEDED_PREFIX)


def prefetch_github_data(row_inputs: list[RowInput], github: Any, github_workers: int) -> GitHubDataCache:
    issue_urls = unique_ordered(url for row_input in row_inputs for url in row_input.issue_urls)
    pr_urls = unique_ordered(url for row_input in row_inputs for url in row_input.pr_urls)
    issues = {}
    pulls = {}

    if not issue_urls and not pr_urls:
        return GitHubDataCache(issues=issues, pulls=pulls)

    print(
        f"Prefetching GitHub data: {len(issue_urls)} issue(s), {len(pr_urls)} PR(s) "
        f"with {github_workers} worker(s)",
        flush=True,
    )

    with ThreadPoolExecutor(max_workers=github_workers) as executor:
        futures = {
            executor.submit(github.get_issue, issue_url): ("issue", issue_url)
            for issue_url in issue_urls
        }
        futures.update(
            {
                executor.submit(github.get_pull, pr_url): ("pull", pr_url)
                for pr_url in pr_urls
            }
        )
        for future in as_completed(futures):
            item_type, url = futures[future]
            try:
                data = future.result()
            except Exception as exc:  # noqa: BLE001
                print(f"Failed to prefetch GitHub {item_type} {url}: {exc}", file=sys.stderr, flush=True)
                continue
            if item_type == "issue":
                issues[url] = data
            else:
                pulls[url] = data
    return GitHubDataCache(issues=issues, pulls=pulls)


def generate_note_for_row(
    row_input: RowInput,
    github_cache: GitHubDataCache,
    ai: Any,
) -> RowGenerationResult:
    expected_links = row_input.issue_urls or row_input.pr_urls
    row_context = build_row_context_from_cache(row_input, github_cache)
    contributors = unique_ordered(
        [author for author in row_context.pr_authors if author not in BOT_AUTHORS]
    )
    try:
        prompt = build_generation_prompt(row_context, expected_links, contributors)
        generated = ai.generate(prompt, expected_links, contributors)
        return RowGenerationResult(
            row_number=row_input.row_number,
            component=row_input.component,
            raw_component=row_input.raw_component,
            note_type=generated.note_type,
            note=generated.release_note,
            error=None,
            needs_review=generated.needs_review,
            reason=generated.reason,
        )
    except Exception as exc:  # noqa: BLE001
        return RowGenerationResult(
            row_number=row_input.row_number,
            component=row_input.component,
            raw_component=row_input.raw_component,
            note_type=None,
            note=None,
            error=str(exc),
        )


def build_row_context_from_cache(row_input: RowInput, github_cache: GitHubDataCache) -> RowContext:
    pr_authors = list(row_input.pr_authors)
    issues = [
        github_cache.issues[issue_url]
        for issue_url in row_input.issue_urls
        if issue_url in github_cache.issues
    ]
    pulls = []
    for pr_url in row_input.pr_urls:
        pull = github_cache.pulls.get(pr_url)
        if not pull:
            continue
        pulls.append(pull)
        if pull.author:
            pr_authors.append(pull.author)
    return RowContext(
        row_number=row_input.row_number,
        component=row_input.component,
        raw_component=row_input.raw_component,
        issue_type=row_input.issue_type,
        pr_title=row_input.pr_title,
        pr_authors=unique_ordered(pr_authors),
        pr_urls=row_input.pr_urls,
        issue_urls=row_input.issue_urls,
        formatted_release_note=row_input.formatted_release_note,
        issues=issues,
        pulls=pulls,
    )


def apply_generation_result(
    sheet: Any,
    header: dict[str, int],
    result: RowGenerationResult,
    entries_by_row: dict[int, list[MarkdownEntry]],
) -> None:
    ai_cell = sheet.cell(row=result.row_number, column=header["release_notes_written_by_ai"])
    if result.error:
        ai_cell.value = f"AI_GENERATION_FAILED: {result.error}"
        print(
            f"Row {result.row_number}: AI generation failed: {result.error}",
            file=sys.stderr,
            flush=True,
        )
        return
    if not result.note or not result.note_type:
        ai_cell.value = "AI_GENERATION_FAILED: empty AI generation result"
        print(
            f"Row {result.row_number}: AI generation failed: empty AI generation result",
            file=sys.stderr,
            flush=True,
        )
        return

    if result.note_type == "not_needed":
        ai_cell.value = result.note
        print(
            f"Row {result.row_number}: {result.note}",
            flush=True,
        )
        return

    ai_cell.value = result.note
    entries_by_row[result.row_number] = [
        MarkdownEntry(result.note_type, result.component, result.note, result.raw_component)
    ]
    review_marker = " (needs review)" if result.needs_review else ""
    print(
        f"Row {result.row_number}: generated {result.note_type}{review_marker}: {result.reason}",
        flush=True,
    )


def release_component_for_row(sheet: Any, header: dict[str, int], row_number: int) -> str:
    raw_component = normalize_raw_component(
        sheet.cell(row=row_number, column=get_component_col(header)).value
    )
    raw_lower = raw_component.lower()
    raw_release_component = release_component_from_raw(raw_component)
    if raw_release_component:
        return raw_release_component

    urls = issue_urls_for_row(sheet, header, row_number)
    urls.extend(extract_pr_urls(str_value(sheet.cell(row=row_number, column=header["pr_link"]).value)))
    repos = {match.group("repo").lower() for url in urls for match in [GITHUB_ITEM_URL_RE.search(url)] if match}

    if "pd" in repos:
        return "PD"
    if "tikv" in repos:
        return "TiKV"
    if "tiflash" in repos:
        return "TiFlash"
    if "ng-monitoring" in repos:
        return "TiDB"
    if "tiup" in repos:
        return "TiUP"
    if repos.intersection({"tiflow", "ticdc"}):
        if "dm" in raw_lower and "cdc" not in raw_lower:
            return "TiDB Data Migration (DM)"
        return "TiCDC"
    if "tidb" in repos:
        if "br" in raw_lower:
            return "Backup & Restore (BR)"
        if "lightning" in raw_lower:
            return "TiDB Lightning"
        if "dumpling" in raw_lower:
            return "Dumpling"
        return "TiDB"
    if "tidb-dashboard" in repos:
        return "TiDB"
    return normalize_component(raw_component)


def release_component_from_raw(raw_component: str) -> str:
    normalized_raw = normalize_component(raw_component)
    if normalized_raw in TOP_LEVEL_COMPONENTS or normalized_raw in TOOL_COMPONENTS:
        return normalized_raw

    token_components = [
        normalize_component(token)
        for token in split_multi_value(raw_component)
    ]
    if not token_components:
        return ""

    for component in [
        "Backup & Restore (BR)",
        "TiDB Lightning",
        "Dumpling",
        "TiUP",
        "sync-diff-inspector",
    ]:
        if component in token_components:
            return component

    for component in TOP_LEVEL_COMPONENTS:
        if component in token_components:
            return component

    if "TiDB Data Migration (DM)" in token_components:
        return "TiDB Data Migration (DM)"
    if "TiCDC" in token_components:
        return "TiCDC"

    return ""


def classify_note_level(note_level: str) -> tuple[str | None, str | None]:
    labels = [label.strip() for label in re.findall(r">\s*([^>]+)", note_level)]
    if not labels:
        return None, None
    section = labels[0].lower()
    note_type = None
    if "bug fixes" in section or "error fixes" in section:
        note_type = "bug_fix"
    elif "improvements" in section:
        note_type = "improvement"

    component_labels = labels[1:]
    if component_labels and component_labels[0].lower() == "tools":
        component_labels = component_labels[1:]
    for label in reversed(component_labels):
        component = normalized_release_component(label)
        if component:
            return note_type, component
    return note_type, None


def classify_note_type_from_text(note: str, issue_type: str) -> str | None:
    note_lower = note.lower()
    issue_type_lower = issue_type.lower()
    if "> bug fixes" in note_lower or "> 错误修复" in note_lower:
        return "bug_fix"
    if "> improvements" in note_lower or "> 改进提升" in note_lower:
        return "improvement"
    if "bug" in issue_type_lower or "fix" in issue_type_lower:
        return "bug_fix"
    if "improvement" in issue_type_lower or "enhancement" in issue_type_lower:
        return "improvement"
    if note.strip().startswith("- Fix "):
        return "bug_fix"
    return "improvement"


def parse_component_from_dup(note: str) -> str | None:
    labels = [label.strip() for label in re.findall(r">\s*([^>]+)", note)]
    cleaned: list[str] = []
    for label in labels:
        if " - " in label:
            label = label.split(" - ", 1)[0]
        cleaned.append(label.strip())
    if len(cleaned) < 2:
        return None
    return normalized_release_component(cleaned[-1])


def fill_row(sheet: Any, row_number: int) -> None:
    for column in range(1, sheet.max_column + 1):
        sheet.cell(row=row_number, column=column).fill = copy.copy(GRAY_FILL)
