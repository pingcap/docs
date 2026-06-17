from __future__ import annotations

import argparse
import os
import tempfile
from pathlib import Path

import openpyxl

from .ai_client import AzureOpenAIClient, CodexAIClient
from .excel_workbook import (
    clear_output_columns,
    collect_markdown_entries_from_sheet,
    generate_notes_without_ai,
    generate_notes_for_sheet,
    get_header,
    merge_rows_by_issue_and_component,
    move_not_needed_rows_to_sheet,
    move_rows_with_issues_already_in_same_series,
    prepare_sheet_columns,
    sort_sheet_rows_by_component,
    store_existing_release_notes,
    update_pr_authors_and_dup_notes,
)
from .github_client import GitHubClient
from .markdown_writer import write_release_file
from .scope_filter import move_prs_not_in_scope, parse_date_value


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Generate English release notes with AI according to PRs and issues "
            "in a specified Excel file. Use subcommands 'generate' and 'export-markdown' "
            "to run the two phases independently."
        ),
    )
    subparsers = parser.add_subparsers(dest="command")

    # --- Phase 1: generate ---
    gen_parser = subparsers.add_parser(
        "generate",
        help=(
            "Phase 1: Process the Excel workbook — run preprocessing, call AI to "
            "generate release notes, and write results back to Excel. "
            "Does NOT produce a Markdown file."
        ),
    )
    add_generate_args(gen_parser)

    # --- Phase 2: export-markdown ---
    export_parser = subparsers.add_parser(
        "export-markdown",
        help=(
            "Phase 2: Read a processed Excel workbook and export a Markdown "
            "release-note file. Does NOT call AI or modify the Excel."
        ),
    )
    add_export_markdown_args(export_parser)

    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        raise SystemExit(1)
    return args


def add_generate_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--version", required=True, help="Target TiDB version, for example 8.5.7.")
    parser.add_argument("--excel", required=True, help="Path to the release note Excel workbook.")
    parser.add_argument(
        "--releases-dir",
        required=True,
        help="Path to the existing English release notes directory.",
    )
    parser.add_argument("--sheet", default="pr_for_release_note", help="Workbook sheet name.")
    parser.add_argument(
        "--ai-provider",
        choices=["codex", "azure"],
        default="codex",
        help=(
            "AI provider to use. 'codex' runs the Codex CLI as a subprocess "
            "(requires codex to be installed). 'azure' calls Azure OpenAI via the "
            "OpenAI Python SDK (requires AZURE_OPENAI_KEY and AZURE_OPENAI_BASE_URL "
            "or OPENAI_BASE_URL environment variables). Default: codex."
        ),
    )
    parser.add_argument(
        "--ai-command",
        default="codex --ask-for-approval never exec --sandbox read-only --ephemeral",
        help="Command-line AI command (only used with --ai-provider codex). The prompt is passed through stdin.",
    )
    parser.add_argument(
        "--ai-model",
        default="gpt-5.4",
        help="Model name. Passed to codex exec with -m, or used as the model parameter for Azure OpenAI.",
    )
    parser.add_argument(
        "--involve-ai-generation",
        type=parse_on_off,
        default="ON",
        help=(
            "Whether to use AI for non-dup release notes. Use ON to generate with AI, "
            "or OFF to skip AI generation and only run preprocessing. Default: ON."
        ),
    )
    parser.add_argument(
        "--ai-timeout",
        type=int,
        default=600,
        help="Timeout in seconds for each AI command invocation.",
    )
    parser.add_argument(
        "--ai-workers",
        type=int,
        default=3,
        help=(
            "Number of concurrent AI command invocations. The default is conservative "
            "for codex exec subprocesses."
        ),
    )
    parser.add_argument(
        "--github-workers",
        type=int,
        default=8,
        help="Number of concurrent GitHub API prefetch workers.",
    )
    parser.add_argument(
        "--author-workers",
        type=int,
        default=3,
        help="Number of concurrent workers used to resolve bot-authored cherry-pick PR authors.",
    )
    parser.add_argument(
        "--checkpoint-interval",
        type=int,
        default=1,
        help=(
            "Save the Excel workbook after every N completed AI rows. "
            "Default: 1. Use 0 to disable."
        ),
    )
    parser.add_argument(
        "--force-regenerate",
        action="store_true",
        help="Clear existing AI release notes and regenerate all non-dup rows.",
    )
    parser.add_argument(
        "--skip-scope-preprocess",
        action="store_true",
        help="Skip moving not-in-scope PR rows to the PRs_not_in_scope sheet.",
    )
    parser.add_argument(
        "--scope-base-branch-start-date",
        help=(
            "Override the estimated release-m.n branch start date for x.y.0 scope "
            "preprocessing, in YYYY-MM-DD format."
        ),
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=None,
        help=(
            "Excel row number to start processing from (1-indexed, row 1 is the header). "
            "Use this to resume from a previous interruption. When specified, "
            "preprocessing steps (sort, merge, scope filter, same-series move) are "
            "skipped because they should have been completed in the first run. "
            "Default: process all data rows."
        ),
    )
    parser.add_argument(
        "--end-row",
        type=int,
        default=None,
        help=(
            "Excel row number to stop processing at (inclusive, 1-indexed). "
            "Default: last row in the sheet."
        ),
    )
    parser.add_argument(
        "--output-excel",
        default=None,
        help=(
            "Path for the processed Excel output. "
            "Default: <original-stem>_processed.xlsx in the same directory."
        ),
    )


def add_export_markdown_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--version", required=True, help="Target TiDB version, for example 8.5.7.")
    parser.add_argument(
        "--excel",
        required=True,
        help="Path to the processed Excel workbook (output of the 'generate' phase).",
    )
    parser.add_argument("--sheet", default="pr_for_release_note", help="Workbook sheet name.")
    parser.add_argument(
        "--releases-dir",
        required=True,
        help="Path to the existing English release notes directory (used for default output path).",
    )
    parser.add_argument(
        "--output-release-file",
        help=(
            "Output Markdown file. Defaults to release-{version}-updated-by-ai.md "
            "if release-{version}.md already exists, otherwise release-{version}.md."
        ),
    )
    parser.add_argument(
        "--release-date",
        default="TBD",
        help='Release date text for the Markdown header, for example "August 14, 2025".',
    )


def main() -> int:
    args = parse_args()
    if args.command == "generate":
        return run_generate(args)
    if args.command == "export-markdown":
        return run_export_markdown(args)
    return 1


def run_generate(args: argparse.Namespace) -> int:
    validate_positive_int("--ai-workers", args.ai_workers)
    validate_positive_int("--github-workers", args.github_workers)
    validate_positive_int("--author-workers", args.author_workers)
    if args.checkpoint_interval < 0:
        raise ValueError("--checkpoint-interval must be greater than or equal to 0")
    base_branch_start_date = None
    if args.scope_base_branch_start_date:
        base_branch_start_date = parse_date_value(args.scope_base_branch_start_date)
        if not base_branch_start_date:
            raise ValueError("--scope-base-branch-start-date must use YYYY-MM-DD format")

    row_range_specified = args.start_row is not None or args.end_row is not None
    start_row = args.start_row
    end_row = args.end_row
    if start_row is not None and start_row < 2:
        raise ValueError("--start-row must be >= 2 (row 1 is the header)")
    if end_row is not None and end_row < 2:
        raise ValueError("--end-row must be >= 2 (row 1 is the header)")
    if start_row is not None and end_row is not None and start_row > end_row:
        raise ValueError("--start-row must be <= --end-row")

    try:
        token = load_github_token()
    except ValueError as exc:
        raise SystemExit(f"error: {exc}") from None
    github = GitHubClient(token)
    involve_ai_generation = args.involve_ai_generation == "ON"
    if involve_ai_generation:
        if args.ai_provider == "azure":
            ai = AzureOpenAIClient(args.ai_model, args.ai_timeout)
        else:
            ai = CodexAIClient(args.ai_command, args.ai_model, args.ai_timeout)
    else:
        ai = None

    excel_path = Path(args.excel)
    processed_excel_path = (
        Path(args.output_excel) if args.output_excel
        else default_processed_excel_path(excel_path)
    )
    workbook = openpyxl.load_workbook(excel_path)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"Cannot find sheet {args.sheet!r} in {args.excel}")
    sheet = workbook[args.sheet]

    if row_range_specified:
        print(
            f"Row range specified: processing rows "
            f"{start_row or 2} to {end_row or sheet.max_row} "
            f"(skipping preprocessing steps)",
            flush=True,
        )
        header = prepare_sheet_columns(sheet)
        if args.force_regenerate:
            clear_output_columns(
                sheet, header, clear_ai=True,
                start_row=start_row, end_row=end_row,
            )
    else:
        if not args.skip_scope_preprocess:
            move_prs_not_in_scope(
                workbook,
                sheet,
                args.version,
                Path(args.releases_dir),
                github,
                base_branch_start_date=base_branch_start_date,
            )
        sort_sheet_rows_by_component(sheet)
        header = prepare_sheet_columns(sheet)
        clear_output_columns(sheet, header, clear_ai=args.force_regenerate)

        existing_notes = store_existing_release_notes(Path(args.releases_dir), args.version)
        move_rows_with_issues_already_in_same_series(
            workbook,
            sheet,
            header,
            existing_notes,
            args.version,
        )
        update_pr_authors_and_dup_notes(
            sheet,
            header,
            existing_notes,
            github,
            author_workers=args.author_workers,
        )
        merge_rows_by_issue_and_component(sheet, header)

    if involve_ai_generation:
        checkpoint_callback = build_checkpoint_callback(
            workbook,
            processed_excel_path,
            args.checkpoint_interval,
        )
        generate_notes_for_sheet(
            sheet,
            header,
            github,
            ai,
            ai_workers=args.ai_workers,
            github_workers=args.github_workers,
            checkpoint_callback=checkpoint_callback,
            start_row=start_row,
            end_row=end_row,
        )
    else:
        generate_notes_without_ai(
            sheet, header, start_row=start_row, end_row=end_row,
        )

    move_not_needed_rows_to_sheet(
        workbook, sheet, header, start_row=start_row, end_row=end_row,
    )
    save_workbook_safely(workbook, processed_excel_path)

    print(f"Phase 1 (generate) completed.", flush=True)
    print(f"  Input Excel: {excel_path}", flush=True)
    print(f"  Processed Excel: {processed_excel_path}", flush=True)
    print(
        f"  Next step: run 'export-markdown' with --excel {processed_excel_path} "
        f"to generate the Markdown file.",
        flush=True,
    )
    return 0


def run_export_markdown(args: argparse.Namespace) -> int:
    excel_path = Path(args.excel)
    output_file = (
        Path(args.output_release_file)
        if args.output_release_file
        else default_output_release_file(Path(args.releases_dir), args.version)
    )

    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"Cannot find sheet {args.sheet!r} in {args.excel}")
    sheet = workbook[args.sheet]
    header = get_header(sheet)

    if "release_notes_written_by_ai" not in header:
        raise ValueError(
            f"Sheet {args.sheet!r} does not have a 'release_notes_written_by_ai' column. "
            "Make sure you are pointing to the processed Excel from the 'generate' phase."
        )

    markdown_entries = collect_markdown_entries_from_sheet(sheet, header)
    workbook.close()

    write_release_file(output_file, args.version, args.release_date, markdown_entries)

    print(f"Phase 2 (export-markdown) completed.", flush=True)
    print(f"  Input Excel: {excel_path}", flush=True)
    print(f"  Generated release note file: {output_file}", flush=True)
    return 0


def validate_positive_int(name: str, value: int) -> None:
    if value < 1:
        raise ValueError(f"{name} must be greater than or equal to 1")


def parse_on_off(value: str) -> str:
    normalized = value.strip().upper()
    if normalized not in {"ON", "OFF"}:
        raise argparse.ArgumentTypeError("value must be ON or OFF")
    return normalized


def default_output_release_file(releases_dir: Path, version: str) -> Path:
    release_file = releases_dir / f"release-{version}.md"
    if release_file.is_file():
        return releases_dir / f"release-{version}-updated-by-ai.md"
    return release_file


def default_processed_excel_path(excel_path: Path) -> Path:
    stem = excel_path.stem
    if stem.endswith("_processed"):
        return excel_path
    return excel_path.with_name(f"{stem}_processed{excel_path.suffix}")


def build_checkpoint_callback(
    workbook: openpyxl.Workbook,
    excel_path: Path,
    checkpoint_interval: int,
):
    if checkpoint_interval <= 0:
        return None

    def checkpoint(completed: int, total: int) -> None:
        if completed % checkpoint_interval != 0 and completed != total:
            return
        save_workbook_safely(workbook, excel_path)
        print(
            f"Checkpoint saved after {completed}/{total} AI row(s): {excel_path}",
            flush=True,
        )

    return checkpoint


def save_workbook_safely(workbook: openpyxl.Workbook, excel_path: Path) -> None:
    excel_path = excel_path.resolve()
    temp_file = tempfile.NamedTemporaryFile(
        prefix=f".{excel_path.stem}.",
        suffix=excel_path.suffix,
        dir=excel_path.parent,
        delete=False,
    )
    temp_path = Path(temp_file.name)
    temp_file.close()
    saved_temp = False
    try:
        workbook.save(temp_path)
        saved_temp = True
        os.replace(temp_path, excel_path)
    except Exception as exc:
        if saved_temp and temp_path.exists():
            raise RuntimeError(
                f"Failed to replace {excel_path}: {exc}. "
                f"A complete temporary workbook remains at {temp_path}."
            ) from exc
        temp_path.unlink(missing_ok=True)
        raise RuntimeError(f"Failed to save workbook {excel_path}: {exc}") from exc


def load_github_token() -> str:
    token = os.environ.get("GITHUB_TOKEN", "").strip()
    if not token:
        raise ValueError("GITHUB_TOKEN environment variable is required")
    return token
